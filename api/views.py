import django_filters
import datetime
from typing import Optional
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_headers, vary_on_cookie
from django.core.cache import cache
from utils.cache import smart_cache_get, smart_cache_set, smart_cache_delete
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
    extend_schema_view,
    inline_serializer,
)
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser, IsAuthenticatedOrReadOnly
from openpyxl import load_workbook
from client.models import Client, ClientImage
from nomenklatura.models import Nomenklatura, NomenklaturaImage
from utils.excel import (
    build_template_workbook,
    clean_cell,
    parse_bool_cell,
    workbook_to_response,
)
from .models import Project, ProjectImage, ImageStatus, ImageSource, AgentLocation
from .serializers import (
    ProjectImageBulkUploadSerializer,
    ProjectImageSerializer,
    ProjectSerializer,
    ProjectDetailSerializer,
    ImageStatusSerializer,
    ImageSourceSerializer,
    ThumbnailEntrySerializer,
    AgentLocationSerializer,
)


class ProjectFilterSet(django_filters.FilterSet):
    description_status = django_filters.ChoiceFilter(
        label="Description status",
        method="filter_description",
        choices=(("with", "Description bor"), ("without", "Description yo'q")),
    )
    image_status = django_filters.ChoiceFilter(
        label="Image status",
        method="filter_image_status",
        choices=(("with", "Rasm bor"), ("without", "Rasm yo'q")),
    )
    created_from = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    created_to = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')
    updated_from = django_filters.DateFilter(field_name='updated_at', lookup_expr='date__gte')
    updated_to = django_filters.DateFilter(field_name='updated_at', lookup_expr='date__lte')

    class Meta:
        model = Project
        fields = ['code_1c', 'name', 'description_status', 'image_status', 'created_from', 'created_to', 'updated_from', 'updated_to', 'is_active']

    def filter_description(self, queryset, name, value):
        if value == "with":
            return queryset.exclude(Q(description__isnull=True) | Q(description__exact=""))
        if value == "without":
            return queryset.filter(Q(description__isnull=True) | Q(description__exact=""))
        return queryset

    def filter_image_status(self, queryset, name, value):
        if value == "with":
            # Rasmlari bor bo'lganlar
            return queryset.filter(images__is_deleted=False).distinct()
        if value == "without":
            # Rasmlari yo'q bo'lganlar
            return queryset.exclude(images__is_deleted=False).distinct()
        return queryset


class ProjectImageFilterSet(django_filters.FilterSet):
    project = django_filters.CharFilter(field_name='project__code_1c')
    project_id = django_filters.NumberFilter(field_name='project__id')
    created_from = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    created_to = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')

    class Meta:
        model = ProjectImage
        fields = ['project', 'project_id', 'is_main', 'category', 'created_from', 'created_to']


@extend_schema_view(
    list=extend_schema(
        tags=['Projects'],
        summary="Project ro'yxatini olish",
        description=(
            "Aktiv va soft-delete qilinmagan (is_deleted=False) project'lar ro'yxatini pagination bilan qaytaradi. "
            "Ma'lumotlar prefetch_related orqali optimallashtirilgan va 5 daqiqaga keshlanadi. "
            "`search` parametri orqali nom, title va `code_1c` bo'yicha qidirish mumkin."
        ),
        parameters=[
            OpenApiParameter(
                name='search',
                required=False,
                type=OpenApiTypes.STR,
                description="Project nomi, title yoki code_1c bo'yicha qidirish",
            ),
            OpenApiParameter(
                name='page',
                required=False,
                type=OpenApiTypes.INT,
                description="Sahifa raqami (default: 1)",
            ),
            OpenApiParameter(
                name='page_size',
                required=False,
                type=OpenApiTypes.INT,
                description="Sahifadagi elementlar soni (default: 20, max: 100)",
            ),
            OpenApiParameter(
                name='code_1c',
                required=False,
                type=OpenApiTypes.STR,
                description="Aniq code_1c bo'yicha filter",
            ),
            OpenApiParameter(
                name='name',
                required=False,
                type=OpenApiTypes.STR,
                description="Aniq nom bo'yicha filter",
            ),
            OpenApiParameter(
                name='description_status',
                required=False,
                type=OpenApiTypes.STR,
                description="Description bo'yicha filter (`with` | `without`)",
            ),
            OpenApiParameter(
                name='image_status',
                required=False,
                type=OpenApiTypes.STR,
                description="Rasm holati bo'yicha filter (`with` - rasm bor | `without` - rasm yo'q)",
            ),
            OpenApiParameter(
                name='created_from',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sanadan boshlab (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name='created_to',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sana chegarasi (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name='updated_from',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yangilangan sanadan boshlab (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name='updated_to',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yangilangan sana chegarasi (YYYY-MM-DD)",
            ),
        ],
    ),
    retrieve=extend_schema(
        tags=['Projects'],
        summary="Bitta project ma'lumotini olish",
        description="`code_1c` identifikatoriga ko'ra project ma'lumotlarini qaytaradi. Ma'lumotlar 10 daqiqaga keshlanadi.",
    ),
    create=extend_schema(
        tags=['Projects'],
        summary="Project yaratish",
        description="Yangi project yozuvini yaratadi. `code_1c` unikal bo'lishi kerak. Yaratilgandan so'ng kesh avtomatik tozalanadi.",
    ),
    update=extend_schema(
        tags=['Projects'],
        summary="Project ma'lumotlarini to'liq yangilash",
        description="`PUT` so'rovi butun project yozuvini yangilaydi.",
    ),
    partial_update=extend_schema(
        tags=['Projects'],
        summary="Project ma'lumotlarini qisman yangilash",
        description="Faqat yuborilgan maydonlarni yangilaydi. `code_1c` o'zgarmaydi.",
    ),
    destroy=extend_schema(
        tags=['Projects'],
        summary="Project'ni soft-delete qilish",
        description="Yozuvni o'chirmasdan, `is_deleted=True` qilib belgilaydi va keshdan o'chiradi.",
        responses={204: OpenApiResponse(description="Project soft-delete qilindi")},
    ),
)
class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.filter(is_deleted=False)
    serializer_class = ProjectSerializer
    lookup_field = 'code_1c'
    lookup_value_regex = '.+'
    filterset_class = ProjectFilterSet
    search_fields = ['code_1c', 'name', 'title']
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ProjectDetailSerializer
        return ProjectSerializer
    
    def get_queryset(self):
        """Optimizatsiya: prefetch_related bilan images yuklash - N+1 query muammosini hal qiladi"""
        # Cache key based on filters
        cache_key = f"project_queryset_{hash(str(self.request.query_params))}"
        cached_qs = smart_cache_get(cache_key)
        if cached_qs is None:
            qs = Project.objects.filter(
                is_deleted=False
            ).prefetch_related(
                'images',
                'images__status',
                'images__source'
            ).order_by('-created_at')
            smart_cache_set(cache_key, qs, 300)  # Cache for 5 minutes
            return qs
        return cached_qs
    
    @method_decorator(cache_page(300))  # Cache list view for 5 minutes
    @method_decorator(vary_on_headers('Authorization'))
    def list(self, request, *args, **kwargs):
        """Cached list view"""
        return super().list(request, *args, **kwargs)
    
    @method_decorator(cache_page(600))  # Cache detail view for 10 minutes
    def retrieve(self, request, *args, **kwargs):
        """Cached detail view"""
        return super().retrieve(request, *args, **kwargs)
    
    def perform_create(self, serializer):
        """Invalidate cache on create"""
        super().perform_create(serializer)
        # Barcha keshni tozalash o'rniga smart_cache_delete ishlatish mumkin, 
        # lekin hozircha cache.clear() osonroq (fallback ni ham tozalash kerak)
        cache.clear()
        from django.core.cache import caches
        caches['fallback'].clear()
    
    def perform_update(self, serializer):
        """Invalidate cache on update"""
        super().perform_update(serializer)
        cache.clear()
        from django.core.cache import caches
        caches['fallback'].clear()
    
    def perform_destroy(self, instance):
        """Soft delete - projectni o'chirmasdan is_deleted=True qiladi"""
        instance.is_deleted = True
        instance.save(update_fields=['is_deleted', 'updated_at'])
        cache.clear()
        from django.core.cache import caches
        caches['fallback'].clear()
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def perform_destroy(self, instance):
        """Soft delete - projectni o'chirmasdan is_deleted=True qiladi"""
        instance.is_deleted = True
        instance.save(update_fields=['is_deleted', 'updated_at'])

    # Excel helpers ---------------------------------------------------------
    @staticmethod
    def _project_excel_headers():
        return ['code_1c', 'name', 'title', 'description', 'is_active']

    def _validate_project_headers(self, sheet):
        expected = [header.lower() for header in self._project_excel_headers()]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(expected), values_only=True), None)
        if not header_row:
            return False, expected, []
        normalized = [clean_cell(value).lower() for value in header_row[:len(expected)]]
        return normalized == expected, expected, normalized

    @extend_schema(
        tags=['Projects'],
        summary="Project ma'lumotlarini Excel formatida eksport qilish",
        description="Filtrlangan project ro'yxatini XLSX fayl sifatida qaytaradi.",
        responses={200: OpenApiResponse(description="XLSX fayl")},
    )
    @action(detail=False, methods=['get'], url_path='export-xlsx', permission_classes=[IsAuthenticated])
    def export_xlsx(self, request):
        # Export uchun alohida queryset - prefetch_related kerak emas va SQLite limit muammosini oldini oladi
        base_queryset = Project.objects.filter(is_deleted=False)
        queryset = self.filter_queryset(base_queryset)
        
        workbook = build_template_workbook('Projects', self._project_excel_headers())
        sheet = workbook.active
        
        # iterator() ishlatish - memory-efficient va SQLite limit muammosini hal qiladi
        # Chunking bilan ishlash - har safar 1000 ta yozuvni qayta ishlash
        chunk_size = 1000
        offset = 0
        
        while True:
            chunk = queryset[offset:offset + chunk_size]
            if not chunk.exists():
                break
            
            for project in chunk.iterator():
                sheet.append([
                    project.code_1c,
                    project.name,
                    project.title or '',
                    project.description or '',
                    project.is_active,
                ])
            
            offset += chunk_size
        
        return workbook_to_response(workbook, 'projects.xlsx')

    @extend_schema(
        tags=['Projects'],
        summary="Project Excel shablonini yuklab olish",
        responses={200: OpenApiResponse(description="XLSX shablon fayl")},
    )
    @action(detail=False, methods=['get'], url_path='template-xlsx', permission_classes=[IsAuthenticated])
    def template_xlsx(self, request):
        workbook = build_template_workbook(
            'Projects template',
            self._project_excel_headers(),
            ['PRJ-001', 'Namuna project', 'Namuna sarlavha', '<p>Namuna description</p>', True],
        )
        return workbook_to_response(workbook, 'projects_template.xlsx')

    @extend_schema(
        tags=['Projects'],
        summary="Project ma'lumotlarini Excel fayldan import qilish",
        request={
            'multipart/form-data': inline_serializer(
                name='ProjectImportPayload',
                fields={
                    'file': serializers.FileField(help_text='XLSX fayl'),
                },
            )
        },
        responses={
            200: OpenApiResponse(description="Import natijalari (created/updated/errors)"),
            400: OpenApiResponse(description="Yaroqsiz fayl yoki header mos emas"),
        },
    )
    @action(
        detail=False,
        methods=['post'],
        url_path='import-xlsx',
        parser_classes=[MultiPartParser],
        permission_classes=[IsAuthenticated],
    )
    def import_xlsx(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'file field talab qilinadi'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(uploaded, data_only=True)
        except Exception:
            return Response({'error': 'XLSX faylni o\'qib bo\'lmadi'}, status=status.HTTP_400_BAD_REQUEST)

        sheet = workbook.active
        is_valid, expected, received = self._validate_project_headers(sheet)
        if not is_valid:
            return Response(
                {
                    'error': 'Excel headerlari mos emas',
                    'expected': expected,
                    'received': received,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        stats = {'created': 0, 'updated': 0, 'errors': []}
        for idx, row in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(expected), values_only=True),
            start=2,
        ):
            if not row or all(value in (None, '') for value in row):
                continue
            code = clean_cell(row[0])
            if not code:
                stats['errors'].append(f"Row {idx}: code_1c bo'sh bo'lishi mumkin emas")
                continue
            name = clean_cell(row[1]) or code
            title = clean_cell(row[2]) or None
            description = row[3] if row[3] is not None else ''
            is_active = parse_bool_cell(row[4], default=True)

            defaults = {
                'name': name,
                'title': title,
                'description': description,
                'is_active': is_active,
                'is_deleted': False,
            }
            try:
                obj, created_flag = Project.objects.update_or_create(
                    code_1c=code,
                    defaults=defaults,
                )
                if created_flag:
                    stats['created'] += 1
                else:
                    stats['updated'] += 1
            except Exception as exc:  # noqa: BLE001
                stats['errors'].append(f"Row {idx}: {exc}")
        return Response(stats, status=status.HTTP_200_OK)


class AgentLocationFilterSet(django_filters.FilterSet):
    date_from = django_filters.DateTimeFilter(field_name='created_at', lookup_expr='gte')
    date_to = django_filters.DateTimeFilter(field_name='created_at', lookup_expr='lte')

    class Meta:
        model = AgentLocation
        fields = [
            'agent_code', 'agent_name', 'agent_phone', 'region', 
            'platform', 'device_id', 'device_manufacturer', 'city',
            'date_from', 'date_to'
        ]


@extend_schema_view(
    list=extend_schema(
        tags=['Agent Locations'],
        summary="Agent lokatsiya yozuvlari ro'yxati",
        description=(
            "Mobil agentlar tomonidan yuborilgan barcha geolokatsiya va qurilma ma'lumotlarini qaytaradi. "
            "Filtrlash uchun `agent_code`, `region`, `platform` va sana orqali foydalanish mumkin. "
            "Ma'lumotlar prefetch qilinmasdan, to'g'ridan-to'g'ri bazadan olinadi (high-frequency data)."
        ),
    ),
    create=extend_schema(
        tags=['Agent Locations'],
        summary="Yangi agent lokatsiyasini yaratish",
        description=(
            "Mobil ilova tomonidan yuborilgan kompleks ma'lumotlarni saqlaydi. "
            "Majburiy maydonlar: `agent_code`, `latitude`, `longitude`. "
            "Qo'shimcha ma'lumotlar: qurilma modellari, sensorlar (accelerometer, gyroscope), "
            "batareya holati, tarmoq sifati va xavfsizlik (root/jailbreak) statuslari. "
            "Bu ma'lumotlar keyinchalik tahlil va monitoring uchun xizmat qiladi."
        ),
    ),
    retrieve=extend_schema(
        tags=['Agent Locations'],
        summary="Bitta lokatsiya yozuvi tafsiloti",
        description="Barcha texnik va geografik metama'lumotlar bilan birga yozuvni qaytaradi."
    ),
    destroy=extend_schema(
        tags=['Agent Locations'],
        summary="Lokatsiya yozuvini soft-delete qilish",
        description="Yozuvni `is_deleted=True` qilib belgilaydi."
    ),
)
class AgentLocationViewSet(viewsets.ModelViewSet):
    queryset = AgentLocation.objects.filter(is_deleted=False)
    serializer_class = AgentLocationSerializer
    filterset_class = AgentLocationFilterSet
    permission_classes = [IsAuthenticated]
    search_fields = ['agent_code', 'agent_name', 'agent_phone', 'device_id', 'device_name', 'region']
    ordering = ['-created_at']

    def get_queryset(self):
        return AgentLocation.objects.filter(is_deleted=False).order_by('-created_at')

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.save(update_fields=['is_deleted', 'updated_at'])

    @extend_schema(
        tags=['Agent Locations'],
        summary="Unikal agentlar ro'yxati",
        description="Barcha unikal agentlarning code, name va phone ma'lumotlarini qaytaradi (dropdown uchun)."
    )
    @action(detail=False, methods=['get'], url_path='unique-agents')
    def unique_agents(self, request):
        """Hamma agentlarning unikal ro'yxatini qaytaradi (dropdown uchun)"""
        # SQLite doesn't support .distinct('field').
        # We fetch records and handle distinctness in Python or use a simpler query.
        agents_data = AgentLocation.objects.filter(is_deleted=False).values(
            'agent_code', 'agent_name', 'agent_phone'
        ).order_by('agent_code')
        
        seen = set()
        unique_agents_list = []
        for agent in agents_data:
            if agent['agent_code'] not in seen:
                unique_agents_list.append(agent)
                seen.add(agent['agent_code'])
                
        return Response(unique_agents_list)

    @extend_schema(
        tags=['Agent Locations'],
        summary="Trayektoriya va visitlar",
        description="Agentning ma'lum kundagi harakati va o'sha kundagi visitlarini qaytaradi.",
        parameters=[
            OpenApiParameter(name='agent_code', required=True, type=str),
            OpenApiParameter(name='date', required=True, type=str, description="YYYY-MM-DD")
        ]
    )
    @action(detail=False, methods=['get'], url_path='trajectory')
    def trajectory(self, request):
        """Berilgan agent va sana uchun harakat trayektoriyasini qaytaradi"""
        agent_code = request.query_params.get('agent_code')
        date_str = request.query_params.get('date')  # YYYY-MM-DD
        
        if not agent_code or not date_str:
            return Response(
                {'error': 'agent_code va date parametrlar talab qilinadi'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Bugungi kundagi recordlarni olish
            qs = AgentLocation.objects.filter(
                agent_code=agent_code,
                is_deleted=False,
                created_at__date=date_str
            ).order_by('created_at').only(
                'latitude', 'longitude', 'speed', 'accuracy', 'created_at', 'battery_level'
            )
            
            points = []
            for loc in qs:
                points.append({
                    'lat': float(loc.latitude),
                    'lng': float(loc.longitude),
                    'speed': float(loc.speed) if loc.speed else 0,
                    'acc': float(loc.accuracy) if loc.accuracy else 0,
                    'bat': float(loc.battery_level) if loc.battery_level else 0,
                    'time': loc.created_at.isoformat()
                })
            
            # Shu agent o'sha kuni qilgan visitlarni ham qo'shish (ClientImage)
            # Bu yerda agent_code orqali bog'lash uchun uploader_name ishlatamiz
            from client.models import ClientImage
            
            visits_qs = ClientImage.objects.filter(
                is_deleted=False,
                created_at__date=date_str,
                # Agent name yoki code uploader_name ichida bo'lishi mumkin deb taxmin qilamiz
                source__uploader_name__icontains=agent_code
            ).select_related('client', 'status')
            
            visits = []
            for v in visits_qs:
                # ClientLocation yo'q bo'lsa, rasmni koordinatasi AgentLocation recordlaridan olinishi kerak
                # Lekin rasmda koordinata bo'lmasa, uni ko'rsata olmaymiz.
                # ProjectImage/ClientImage da koordinata maydonlari yo'qligini ko'rdik.
                # Shuning uchun rasm olingan vaqtga eng yaqin AgentLocation ni topib, 
                # o'sha koordinatani rasm koordinatasi deb olamiz.
                
                # Simple approach: Rasm olingan vaqtga eng yaqin nuqtani topamiz
                closest_point = None
                min_diff = float('inf')
                
                v_created_at = v.created_at
                if v_created_at.tzinfo:
                    v_created_at = v_created_at.replace(tzinfo=None)

                for p in points:
                    p_time = datetime.datetime.fromisoformat(p['time'])
                    if p_time.tzinfo:
                        p_time = p_time.replace(tzinfo=None)
                        
                    diff = abs((p_time - v_created_at).total_seconds())
                    if diff < min_diff and diff < 300: # 5 minut ichida bo'lsa
                        min_diff = diff
                        closest_point = p
                
                if closest_point:
                    visits.append({
                        'id': v.id,
                        'client_name': v.client.name,
                        'time': v.created_at.isoformat(),
                        'lat': closest_point['lat'],
                        'lng': closest_point['lng'],
                        'category': v.category,
                        'status': v.status.name if v.status else ''
                    })

            return Response({
                'agent_code': agent_code,
                'date': date_str,
                'points_count': len(points),
                'points': points,
                'visits': visits
            })
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @extend_schema(
        tags=['Agent Locations'],
        summary="Hududiy aktivlik",
        description="Agentning vaqt oralig'idagi regionlar bo'yicha aktivligini qaytaradi.",
        parameters=[
            OpenApiParameter(name='agent_code', required=False, type=str),
            OpenApiParameter(name='date_from', required=False, type=str, description="YYYY-MM-DD"),
            OpenApiParameter(name='date_to', required=False, type=str, description="YYYY-MM-DD")
        ]
    )
    @action(detail=False, methods=['get'], url_path='regional-activity')
    def regional_activity(self, request):
        """Regionlar va shaharlar bo'yicha agent aktivligini qaytaradi"""
        agent_code = request.query_params.get('agent_code')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        qs = AgentLocation.objects.filter(is_deleted=False)
        if agent_code:
            qs = qs.filter(agent_code=agent_code)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Region bo'yicha guruhlash
        region_stats = qs.values('region').annotate(
            points_count=models.Count('id'),
        ).order_by('-points_count')

        # Visitlar bo'yicha (ClientImage)
        # Note: ClientImage modelida source maydoni yo'q bo'lishi mumkin
        total_visits = 0
        try:
            from client.models import ClientImage
            v_qs = ClientImage.objects.filter(is_deleted=False)
            if date_from:
                v_qs = v_qs.filter(created_at__date__gte=date_from)
            if date_to:
                v_qs = v_qs.filter(created_at__date__lte=date_to)
            # Agent code bo'yicha filter qilish qiyin, chunki ClientImage'da to'g'ridan-to'g'ri agent maydoni yo'q
            total_visits = v_qs.count()
        except Exception as e:
            # Agar ClientImage modelida muammo bo'lsa, 0 qaytaramiz
            print(f"Visits count error: {e}")
            total_visits = 0
        
        return Response({
            'period': {
                'from': date_from,
                'to': date_to
            },
            'agent_code': agent_code,
            'regions': list(region_stats),
            'total_points': qs.count(),
            'total_visits': total_visits
        })

@extend_schema_view(
    list=extend_schema(
        tags=['Projects'],
        summary="Project rasmlari ro'yxatini olish",
        description="`project` va `is_main` bo'yicha filterlash mumkin.",
        parameters=[
            OpenApiParameter(
                name='project',
                required=False,
                type=OpenApiTypes.STR,
                description="Project code_1c bo'yicha filter",
            ),
            OpenApiParameter(
                name='is_main',
                required=False,
                type=OpenApiTypes.BOOL,
                description="Asosiy rasm bo'yicha filter",
            ),
            OpenApiParameter(
                name='category',
                required=False,
                type=OpenApiTypes.STR,
                description="Rasm toifasi bo'yicha filter",
            ),
            OpenApiParameter(
                name='created_from',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sanadan boshlab (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name='created_to',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sana chegarasi (YYYY-MM-DD)",
            ),
        ],
    ),
    create=extend_schema(
        tags=['Projects'],
        summary="Project uchun bitta rasm yuklash",
        description="Multipart form-data orqali bitta rasm faylini yuklaydi.",
    ),
    destroy=extend_schema(
        tags=['Projects'],
        summary="Project rasmini o'chirish",
        description="Rasmni soft-delete qiladi yoki bazadan o'chiradi.",
    ),
)
class ProjectImageViewSet(viewsets.ModelViewSet):
    queryset = ProjectImage.objects.filter(is_deleted=False)
    serializer_class = ProjectImageSerializer
    filterset_class = ProjectImageFilterSet
    search_fields = ['project__code_1c', 'project__name']
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        """Optimizatsiya: select_related bilan project yuklash - N+1 query muammosini hal qiladi"""
        return ProjectImage.objects.filter(
            is_deleted=False
        ).select_related('project').order_by('-created_at')
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    @extend_schema(
        tags=['Projects'],
        summary="Project uchun bir nechta rasm yuklash",
        description=(
            "Multipart form-data formatida bir nechta rasmni birdaniga yuklash imkonini beradi. "
            "`project` maydoniga `code_1c` qiymati, `images` maydoniga esa bir yoki bir nechta fayl yuboriladi. "
            "Ixtiyoriy ravishda `category` va `note` maydonlari orqali rasmlar uchun umumiy teg yoki izoh qo'shish mumkin."
        ),
        request=ProjectImageBulkUploadSerializer,
        responses={
            201: OpenApiResponse(
                response=ProjectImageSerializer(many=True),
                description="Yangi yuklangan rasmlar ro'yxati",
            ),
            400: OpenApiResponse(description="Kerakli maydonlar yetishmaydi"),
            404: OpenApiResponse(description="Project topilmadi"),
        },
        examples=[
            OpenApiExample(
                name="Multipart sample",
                description="HTTPie yordamida bir nechta rasm yuborish",
                value="http --form POST /api/v1/project-images/bulk-upload/ project=P-001 "
                "category=marketing note='Banner versiyasi' "
                "images@/path/to/image1.jpg images@/path/to/image2.jpg",
            )
        ],
    )
    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        """Bir vaqtda bir nechta rasm yuklash"""
        project_code = request.data.get('project')
        if not project_code:
            return Response(
                {'error': 'project code_1c talab qilinadi'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            project = Project.objects.get(code_1c=project_code, is_deleted=False)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project topilmadi'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        images = request.FILES.getlist('images')
        if not images:
            return Response(
                {'error': 'Rasmlar talab qilinadi'},
                status=status.HTTP_400_BAD_REQUEST
            )

        category = request.data.get('category', '')
        note = request.data.get('note', '')
        
        created_images = []
        for image in images:
            image_obj = ProjectImage.objects.create(
                project=project,
                image=image,
                is_main=False,
                is_active=True,
                is_deleted=False,
                category=category,
                note=note,
            )
            serializer = ProjectImageSerializer(image_obj, context={'request': request})
            created_images.append(serializer.data)
        
        return Response({
            'message': f'{len(created_images)} ta rasm muvaffaqiyatli yuklandi',
            'images': created_images
        }, status=status.HTTP_201_CREATED)


@extend_schema_view(
    list=extend_schema(
        tags=['Image Management'],
        summary="Image statuslar ro'yxatini olish",
        description="Barcha aktiv image statuslar ro'yxatini qaytaradi. Pagination yo'q.",
    ),
    retrieve=extend_schema(
        tags=['Image Management'],
        summary="Bitta image status ma'lumotini olish",
    ),
    create=extend_schema(
        tags=['Image Management'],
        summary="Yangi image status yaratish",
    ),
    update=extend_schema(
        tags=['Image Management'],
        summary="Image status ma'lumotlarini to'liq yangilash",
    ),
    partial_update=extend_schema(
        tags=['Image Management'],
        summary="Image status ma'lumotlarini qisman yangilash",
    ),
    destroy=extend_schema(
        tags=['Image Management'],
        summary="Image status'ni soft-delete qilish",
    ),
)
class ImageStatusViewSet(viewsets.ModelViewSet):
    """ImageStatus CRUD operatsiyalari - Heavily cached"""
    queryset = ImageStatus.objects.filter(is_deleted=False, is_active=True)
    serializer_class = ImageStatusSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None  # Statuslar kam bo'ladi, pagination kerak emas
    search_fields = ['code', 'name', 'description']
    ordering = ['order', 'name']
    
    @method_decorator(cache_page(3600))  # Cache for 1 hour (rarely changes)
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    @method_decorator(cache_page(3600))
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def perform_create(self, serializer):
        super().perform_create(serializer)
        cache.clear()
    
    def perform_update(self, serializer):
        super().perform_update(serializer)
        cache.clear()
    
    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        cache.clear()


@extend_schema_view(
    list=extend_schema(
        tags=['Image Management'],
        summary="Image source ro'yxatini olish",
        description="Barcha aktiv image source'lar ro'yxatini pagination bilan qaytaradi.",
    ),
    retrieve=extend_schema(
        tags=['Image Management'],
        summary="Bitta image source ma'lumotini olish",
    ),
    create=extend_schema(
        tags=['Image Management'],
        summary="Yangi image source yaratish",
    ),
    update=extend_schema(
        tags=['Image Management'],
        summary="Image source ma'lumotlarini to'liq yangilash",
    ),
    partial_update=extend_schema(
        tags=['Image Management'],
        summary="Image source ma'lumotlarini qisman yangilash",
    ),
    destroy=extend_schema(
        tags=['Image Management'],
        summary="Image source'ni soft-delete qilish",
    ),
)
class ImageSourceViewSet(viewsets.ModelViewSet):
    """ImageSource CRUD operatsiyalari"""
    queryset = ImageSource.objects.filter(is_deleted=False)
    serializer_class = ImageSourceSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ['uploader_name', 'uploader_contact', 'upload_location']
    ordering = ['-created_at']


class ThumbnailFeedMixin:
    """Reusable helper mixin for thumbnail responses"""

    @staticmethod
    def _parse_entity_types(raw: str | None):
        allowed = ['project', 'client', 'nomenklatura']
        if not raw:
            return allowed
        parsed = []
        for part in raw.split(','):
            trimmed = part.strip().lower()
            if trimmed in allowed and trimmed not in parsed:
                parsed.append(trimmed)
        return parsed or allowed

    @staticmethod
    def _parse_limit(raw: str | None) -> int:
        default = 60
        max_limit = 200
        if raw is None:
            return default
        try:
            value = int(raw)
            if value < 1:
                return default
            return min(value, max_limit)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _parse_bool(raw: str | None):
        if raw is None:
            return None
        return raw.strip().lower() in ('1', 'true', 'yes')

    def _collect_project_thumbnails(self, request, limit, is_main, status_code, code_1c=None):
        qs = ProjectImage.objects.filter(
            is_deleted=False,
            is_active=True,
            project__is_deleted=False,
        ).select_related('project', 'status', 'source').order_by('-created_at')
        if is_main is not None:
            qs = qs.filter(is_main=is_main)
        if status_code:
            qs = qs.filter(status__code=status_code)
        if code_1c:
            qs = qs.filter(project__code_1c=code_1c)
        qs = qs[:limit]
        return [
            self._build_entry(
                request,
                entity_type='project',
                entity=image.project,
                image=image,
                code_attr='code_1c'
            )
            for image in qs
        ]

    def _collect_client_thumbnails(self, request, limit, is_main, status_code, client_id=None, client_code_1c=None):
        qs = ClientImage.objects.filter(
            is_deleted=False,
            is_active=True,
            client__is_deleted=False,
        ).select_related('client', 'status', 'source').order_by('-created_at')
        if is_main is not None:
            qs = qs.filter(is_main=is_main)
        if status_code:
            qs = qs.filter(status__code=status_code)
        if client_id:
            qs = qs.filter(client_id=client_id)
        if client_code_1c:
            qs = qs.filter(client__client_code_1c=client_code_1c)
        qs = qs[:limit]
        return [
            self._build_entry(
                request,
                entity_type='client',
                entity=image.client,
                image=image,
                code_attr='client_code_1c'
            )
            for image in qs
        ]

    def _collect_nomenklatura_thumbnails(self, request, limit, is_main, status_code, nomenklatura_id=None, code_1c=None, article_code=None):
        qs = NomenklaturaImage.objects.filter(
            is_deleted=False,
            is_active=True,
            nomenklatura__is_deleted=False,
        ).select_related('nomenklatura', 'status', 'source').order_by('-created_at')
        if is_main is not None:
            qs = qs.filter(is_main=is_main)
        if status_code:
            qs = qs.filter(status__code=status_code)
        if nomenklatura_id:
            qs = qs.filter(nomenklatura_id=nomenklatura_id)
        if code_1c:
            qs = qs.filter(nomenklatura__code_1c=code_1c)
        if article_code:
            qs = qs.filter(nomenklatura__article_code=article_code)
        qs = qs[:limit]
        return [
            self._build_entry(
                request,
                entity_type='nomenklatura',
                entity=image.nomenklatura,
                image=image,
                code_attr='code_1c'
            )
            for image in qs
        ]

    def _get_image_size_bytes(self, image_field) -> Optional[int]:
        """Rasm hajmini bayt'larda olish"""
        if not image_field:
            return None
        try:
            if hasattr(image_field, 'file') and image_field.file:
                return image_field.file.size
        except Exception:
            pass
        return None
    
    def _get_image_size(self, image_field) -> Optional[str]:
        """Rasm hajmini olish (KB yoki MB formatida)"""
        size_bytes = self._get_image_size_bytes(image_field)
        if size_bytes is None:
            return None
        if size_bytes >= 1024 * 1024:  # MB
            return f"{size_bytes / (1024 * 1024):.2f} MB"
        elif size_bytes >= 1024:  # KB
            return f"{size_bytes / 1024:.2f} KB"
        else:  # Bytes
            return f"{size_bytes} B"
    
    def _format_size(self, size_bytes: int) -> str:
        """Bayt'larni KB yoki MB formatiga aylantirish"""
        if size_bytes >= 1024 * 1024:  # MB
            return f"{size_bytes / (1024 * 1024):.2f} MB"
        elif size_bytes >= 1024:  # KB
            return f"{size_bytes / 1024:.2f} KB"
        else:  # Bytes
            return f"{size_bytes} B"
    
    def _get_thumbnail_dimensions(self, image_obj):
        """Thumbnail rasm o'lchamlari va hajmini olish"""
        try:
            thumb = image_obj.image_thumbnail
            if not thumb:
                return None
            size_str = self._get_image_size(thumb)
            return {
                'width': 150,
                'height': 150,
                'format': 'JPEG',
                'size': size_str
            }
        except Exception:
            pass
        return None
    
    def _get_original_dimensions(self, image_obj):
        """Original rasm o'lchamlari va hajmini olish"""
        try:
            if not image_obj.image:
                return None
            from PIL import Image as PILImage
            with PILImage.open(image_obj.image.file) as img:
                file_size = 0
                if hasattr(image_obj.image.file, 'size'):
                    file_size = image_obj.image.file.size
                elif hasattr(image_obj.image, 'size'):
                    file_size = image_obj.image.size
                
                # KB yoki MB ga aylantirish
                if file_size >= 1024 * 1024:  # MB
                    size_str = f"{file_size / (1024 * 1024):.2f} MB"
                elif file_size >= 1024:  # KB
                    size_str = f"{file_size / 1024:.2f} KB"
                else:  # Bytes
                    size_str = f"{file_size} B"
                
                return {
                    'width': img.width,
                    'height': img.height,
                    'format': img.format or 'JPEG',
                    'size_bytes': file_size,
                    'size': size_str
                }
        except Exception:
            pass
        return None

    def _build_entry(self, request, entity_type, entity, image, code_attr):
        status_obj = getattr(image, 'status', None)
        source_obj = getattr(image, 'source', None)
        code_value = getattr(entity, code_attr, None)
        if not code_value:
            code_value = ''
        
        # Rasm o'lchamlari va hajmini olish
        thumbnail_dimensions = self._get_thumbnail_dimensions(image)
        original_dimensions = self._get_original_dimensions(image)
        
        # Thumbnail va original hajmini bayt'larda olish (umumiy hajmni hisoblash uchun)
        thumbnail_size_bytes = None
        if thumbnail_dimensions and 'size' in thumbnail_dimensions:
            # size_str dan bayt'larni olish (agar mavjud bo'lsa)
            try:
                thumb = image.image_thumbnail
                if thumb and hasattr(thumb, 'file') and thumb.file:
                    thumbnail_size_bytes = thumb.file.size
            except Exception:
                pass
        
        original_size_bytes = None
        if original_dimensions and 'size_bytes' in original_dimensions:
            original_size_bytes = original_dimensions['size_bytes']
        
        return {
            'entity_type': entity_type,
            'entity_id': entity.id,
            'code_1c': code_value,
            'entity_name': getattr(entity, 'name', str(entity)),
            'image_id': image.id,
            'thumbnail_url': self._absolute_url(request, image),
            'thumbnail_dimensions': thumbnail_dimensions,
            'original_dimensions': original_dimensions,
            'thumbnail_size_bytes': thumbnail_size_bytes,  # Umumiy hajmni hisoblash uchun
            'original_size_bytes': original_size_bytes,  # Umumiy hajmni hisoblash uchun
            'is_main': image.is_main,
            'category': image.category or None,
            'note': image.note or None,
            'status_code': status_obj.code if status_obj else None,
            'status_name': status_obj.name if status_obj else None,
            'source_name': source_obj.uploader_name if source_obj else None,
            'source_type': source_obj.uploader_type if source_obj else None,
            'created_at': image.created_at,
        }

    @staticmethod
    def _absolute_url(request, image_obj):
        try:
            thumb = image_obj.image_thumbnail
        except Exception:  # noqa: BLE001 - imagekit throws RuntimeError when image absent
            return None
        if not thumb:
            return None
        url = thumb.url
        if not request:
            return url
        return request.build_absolute_uri(url)


@extend_schema(
    tags=['Image Management'],
    summary="Birlashtirilgan thumbnail feed (Project/Client/Nomenklatura)",
    description=(
        "Project, Client va Nomenklatura rasmlarining faqat thumbnail URL larini birlashtirilgan holda qaytaradi. "
        "Har bir element entity haqida batafsil ma'lumotlar (ID, code, nom, o'lchamlar) bilan birga keladi. "
        "Javob tarkibida `total_size` maydoni mavjud bo'lib, u barcha qaytarilgan rasmlarning "
        "umumiy hajmini (thumbnail va original) bayt va inson o'qiydigan formatda ko'rsatadi. "
        "Ushbu endpoint mobil ilovalar uchun preview ro'yxatini tezkor shakllantirishga mo'ljallangan va keshlanadi."
    ),
    parameters=[
        OpenApiParameter(
            name='entity_type',
            type=OpenApiTypes.STR,
            required=False,
            description="Filtrlash uchun entity turlari (virgullar bilan ajratilgan). "
                        "Ruxsat etilgan qiymatlar: project, client, nomenklatura. Default: hammasi.",
        ),
        OpenApiParameter(
            name='is_main',
            type=OpenApiTypes.BOOL,
            required=False,
            description="Faqat asosiy (true) yoki oddiy (false) rasmlarni qaytarish",
        ),
        OpenApiParameter(
            name='status',
            type=OpenApiTypes.STR,
            required=False,
            description="ImageStatus kodi bo'yicha filter (masalan: store_before)",
        ),
        OpenApiParameter(
            name='limit',
            type=OpenApiTypes.INT,
            required=False,
            description="Qaytariladigan maksimal elementlar soni (default: 60, max: 200)",
        ),
    ],
    responses={
        200: inline_serializer(
            name='ThumbnailFeedResponse',
            fields={
                'results': ThumbnailEntrySerializer(many=True),
                'total_count': serializers.IntegerField(),
                'total_size': inline_serializer(
                    name='TotalSizeSummary',
                    fields={
                        'thumbnail': inline_serializer(
                            name='SizeDetail',
                            fields={
                                'size_bytes': serializers.IntegerField(),
                                'size': serializers.CharField(),
                            }
                        ),
                        'original': inline_serializer(
                            name='OriginalSizeDetail',
                            fields={
                                'size_bytes': serializers.IntegerField(),
                                'size': serializers.CharField(),
                            }
                        ),
                    }
                ),
            }
        )
    }
)
@method_decorator(cache_page(180), name='get')  # Cache for 3 minutes
class ThumbnailFeedView(ThumbnailFeedMixin, APIView):
    """Birlashtirilgan thumbnail feed (project/client/nomenklatura) - Cached"""

    permission_classes = [AllowAny]

    def get(self, request):
        # Cache key based on query parameters
        cache_key = f"thumbnail_feed_{hash(str(request.query_params))}"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)
        
        requested_types = self._parse_entity_types(request.query_params.get('entity_type'))
        limit = self._parse_limit(request.query_params.get('limit'))
        is_main = self._parse_bool(request.query_params.get('is_main'))
        status_code = request.query_params.get('status')
        code_1c = request.query_params.get('code_1c')
        client_code_1c = request.query_params.get('client_code_1c') or code_1c
        article_code = request.query_params.get('article_code')

        per_type_limit = max(limit, 1)
        entries = []

        if 'project' in requested_types:
            entries.extend(
                self._collect_project_thumbnails(request, per_type_limit, is_main, status_code, code_1c=code_1c)
            )
        if 'client' in requested_types:
            entries.extend(
                self._collect_client_thumbnails(request, per_type_limit, is_main, status_code, client_code_1c=client_code_1c)
            )
        if 'nomenklatura' in requested_types:
            entries.extend(
                self._collect_nomenklatura_thumbnails(request, per_type_limit, is_main, status_code, code_1c=code_1c, article_code=article_code)
            )

        entries.sort(key=lambda item: item['created_at'], reverse=True)
        final_entries = entries[:limit]
        serialized = ThumbnailEntrySerializer(final_entries, many=True)
        
        # Umumiy hajmni hisoblash
        total_thumbnail_size_bytes = sum(
            entry.get('thumbnail_size_bytes', 0) or 0 
            for entry in final_entries
        )
        total_original_size_bytes = sum(
            entry.get('original_size_bytes', 0) or 0 
            for entry in final_entries
        )
        
        response_data = {
            'results': serialized.data,
            'total_count': len(final_entries),
            'total_size': {
                'thumbnail': {
                    'size_bytes': total_thumbnail_size_bytes,
                    'size': self._format_size(total_thumbnail_size_bytes)
                },
                'original': {
                    'size_bytes': total_original_size_bytes,
                    'size': self._format_size(total_original_size_bytes)
                }
            }
        }
        
        # Cache the response
        cache.set(cache_key, response_data, 180)  # 3 minutes
        
        return Response(response_data, status=status.HTTP_200_OK)


@extend_schema(
    tags=['Image Management'],
    summary="Project thumbnail rasmlari",
    description="Faqat `ProjectImage` larining thumbnail URL larini qaytaradi.",
    parameters=[
        OpenApiParameter(
            name='is_main',
            type=OpenApiTypes.BOOL,
            required=False,
            description="Faqat asosiy (true) yoki oddiy (false) rasmlarni qaytarish",
        ),
        OpenApiParameter(
            name='status',
            type=OpenApiTypes.STR,
            required=False,
            description="ImageStatus kodi bo'yicha filter (masalan: store_before)",
        ),
        OpenApiParameter(
            name='limit',
            type=OpenApiTypes.INT,
            required=False,
            description="Qaytariladigan maksimal elementlar soni (default: 60, max: 200)",
        ),
    ],
    responses={
        200: OpenApiResponse(
            response=ThumbnailEntrySerializer(many=True),
            description="Project thumbnail ma'lumotlari",
        )
    }
)
class ProjectThumbnailView(ThumbnailFeedMixin, APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        limit = self._parse_limit(request.query_params.get('limit'))
        is_main = self._parse_bool(request.query_params.get('is_main'))
        status_code = request.query_params.get('status')
        code_1c = request.query_params.get('code_1c')

        entries = self._collect_project_thumbnails(request, limit, is_main, status_code, code_1c=code_1c)
        serialized = ThumbnailEntrySerializer(entries, many=True)
        
        # Umumiy hajmni hisoblash
        total_thumbnail_size_bytes = sum(
            entry.get('thumbnail_size_bytes', 0) or 0 
            for entry in entries
        )
        total_original_size_bytes = sum(
            entry.get('original_size_bytes', 0) or 0 
            for entry in entries
        )
        
        response_data = {
            'results': serialized.data,
            'total_count': len(entries),
            'total_size': {
                'thumbnail': {
                    'size_bytes': total_thumbnail_size_bytes,
                    'size': self._format_size(total_thumbnail_size_bytes)
                },
                'original': {
                    'size_bytes': total_original_size_bytes,
                    'size': self._format_size(total_original_size_bytes)
                }
            }
        }
        
        return Response(response_data, status=status.HTTP_200_OK)


@extend_schema(
    tags=['Image Management'],
    summary="Client thumbnail rasmlari",
    description="Faqat `ClientImage` larining thumbnail URL larini qaytaradi.",
    parameters=[
        OpenApiParameter(
            name='is_main',
            type=OpenApiTypes.BOOL,
            required=False,
            description="Faqat asosiy (true) yoki oddiy (false) rasmlarni qaytarish",
        ),
        OpenApiParameter(
            name='status',
            type=OpenApiTypes.STR,
            required=False,
            description="ImageStatus kodi bo'yicha filter (masalan: store_before)",
        ),
        OpenApiParameter(
            name='limit',
            type=OpenApiTypes.INT,
            required=False,
            description="Qaytariladigan maksimal elementlar soni (default: 60, max: 200)",
        ),
    ],
    responses={
        200: OpenApiResponse(
            response=ThumbnailEntrySerializer(many=True),
            description="Client thumbnail ma'lumotlari",
        )
    }
)
class ClientThumbnailView(ThumbnailFeedMixin, APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        limit = self._parse_limit(request.query_params.get('limit'))
        is_main = self._parse_bool(request.query_params.get('is_main'))
        status_code = request.query_params.get('status')
        client_id = request.query_params.get('client_id')
        client_code_1c = request.query_params.get('client_code_1c')

        entries = self._collect_client_thumbnails(
            request=request,
            limit=limit,
            is_main=is_main,
            status_code=status_code,
            client_id=client_id,
            client_code_1c=client_code_1c,
        )
        
        # entries = self._collect_client_thumbnails(request, limit, is_main, status_code)
        serialized = ThumbnailEntrySerializer(entries, many=True)
        
        # Umumiy hajmni hisoblash
        total_thumbnail_size_bytes = sum(
            entry.get('thumbnail_size_bytes', 0) or 0 
            for entry in entries
        )
        total_original_size_bytes = sum(
            entry.get('original_size_bytes', 0) or 0 
            for entry in entries
        )
        
        response_data = {
            'results': serialized.data,
            'total_count': len(entries),
            'total_size': {
                'thumbnail': {
                    'size_bytes': total_thumbnail_size_bytes,
                    'size': self._format_size(total_thumbnail_size_bytes)
                },
                'original': {
                    'size_bytes': total_original_size_bytes,
                    'size': self._format_size(total_original_size_bytes)
                }
            }
        }
        
        return Response(response_data, status=status.HTTP_200_OK)


@extend_schema(
    tags=['Image Management'],
    summary="Nomenklatura thumbnail rasmlari",
    description="Faqat `NomenklaturaImage` larining thumbnail URL larini qaytaradi.",
    parameters=[
        OpenApiParameter(
            name='is_main',
            type=OpenApiTypes.BOOL,
            required=False,
            description="Faqat asosiy (true) yoki oddiy (false) rasmlarni qaytarish",
        ),
        OpenApiParameter(
            name='status',
            type=OpenApiTypes.STR,
            required=False,
            description="ImageStatus kodi bo'yicha filter (masalan: store_before)",
        ),
        OpenApiParameter(
            name='limit',
            type=OpenApiTypes.INT,
            required=False,
            description="Qaytariladigan maksimal elementlar soni (default: 60, max: 200)",
        ),
    ],
    responses={
        200: OpenApiResponse(
            response=ThumbnailEntrySerializer(many=True),
            description="Nomenklatura thumbnail ma'lumotlari",
        )
    }
)
class NomenklaturaThumbnailView(ThumbnailFeedMixin, APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        limit = self._parse_limit(request.query_params.get('limit'))
        is_main = self._parse_bool(request.query_params.get('is_main'))
        status_code = request.query_params.get('status')
        code_1c = request.query_params.get('code_1c')
        article_code = request.query_params.get('article_code')

        entries = self._collect_nomenklatura_thumbnails(request, limit, is_main, status_code, code_1c=code_1c, article_code=article_code)
        serialized = ThumbnailEntrySerializer(entries, many=True)
        
        # Umumiy hajmni hisoblash
        total_thumbnail_size_bytes = sum(
            entry.get('thumbnail_size_bytes', 0) or 0 
            for entry in entries
        )
        total_original_size_bytes = sum(
            entry.get('original_size_bytes', 0) or 0 
            for entry in entries
        )
        
        response_data = {
            'results': serialized.data,
            'total_count': len(entries),
            'total_size': {
                'thumbnail': {
                    'size_bytes': total_thumbnail_size_bytes,
                    'size': self._format_size(total_thumbnail_size_bytes)
                },
                'original': {
                    'size_bytes': total_original_size_bytes,
                    'size': self._format_size(total_original_size_bytes)
                }
            }
        }
        
        return Response(response_data, status=status.HTTP_200_OK)

class ClearDatabaseView(APIView):
    """
    Adminlar uchun bazani tozalash messodi.
    Tanlangan table'lardagi barcha ma'lumotlarni o'chirib tashlaydi.
    """
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request):
        # Model mapping
        MODELS_MAP = {
            'client': Client,
            'client_image': ClientImage,
            'nomenklatura': Nomenklatura,
            'nomenklatura_image': NomenklaturaImage,
            'project': Project,
            'project_image': ProjectImage,
            'agent_location': AgentLocation,
            'image_source': ImageSource,
            'image_status': ImageStatus,
        }

        selected_models = request.data.get('models', [])
        if not selected_models:
            return Response(
                {'error': 'Hech qanday table tanlanmadi'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        stats = {}
        errors = []

        # Transaction bilan o'rash tavsiya etiladi, lekin bu "dangerous" operation
        # shuning uchun har biri alohida o'chgani ma'qul bo'lishi mumkin (partial success).
        
        for key in selected_models:
            model = MODELS_MAP.get(key)
            if model:
                try:
                    # cascade delete bo'ladi
                    count, _ = model.objects.all().delete()
                    stats[key] = count
                except Exception as e:
                    errors.append(f"{key}: {str(e)}")
        
        return Response({
            'message': 'Tozalash yakunlandi',
            'stats': stats,
            'errors': errors
        })
