import django_filters
from django.db.models import Q
from nomenklatura.models import Nomenklatura, NomenklaturaImage
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from rest_framework.response import Response
from django.core.cache import cache
from utils.cache import smart_cache_get, smart_cache_set, smart_cache_delete
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
    extend_schema_view,
    inline_serializer,
)
from openpyxl import load_workbook
from utils.excel import build_template_workbook, workbook_to_response, parse_bool_cell, clean_cell
from core.models import ImportLog
from .serializers import (
    NomenklaturaImageBulkUploadSerializer,
    NomenklaturaImageSerializer,
    NomenklaturaSerializer,
)


class NomenklaturaFilterSet(django_filters.FilterSet):
    """
    Enhanced filtering with backward compatibility
    All existing filters preserved, new filters added
    """
    # EXISTING FILTERS (preserved)
    description_status = django_filters.ChoiceFilter(
        label="Description status",
        method="filter_description",
        choices=(("with", "Description bor"), ("without", "Description yo'q")),
    )
    image_status = django_filters.ChoiceFilter(
        label="Image status",
        method="filter_image_status",
        choices=(("with", "Rasm bor"), ("without", "Rasm yo'q")),
    )
    project = django_filters.CharFilter(field_name='project__code_1c', label="Project code_1c")
    project_id = django_filters.NumberFilter(field_name='project__id', label="Project ID")
    created_from = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    created_to = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')
    updated_from = django_filters.DateFilter(field_name='updated_at', lookup_expr='date__gte')
    updated_to = django_filters.DateFilter(field_name='updated_at', lookup_expr='date__lte')
    
    # NEW FILTERS (additive only - no breaking changes)
    search = django_filters.CharFilter(method='filter_search', label="Global search")
    name_contains = django_filters.CharFilter(field_name='name', lookup_expr='icontains', label="Name contains")
    code_contains = django_filters.CharFilter(field_name='code_1c', lookup_expr='icontains', label="Code contains")
    
    # Brand/Manufacturer filters
    brand = django_filters.CharFilter(field_name='brand', lookup_expr='iexact', label="Brand (exact)")
    brand_contains = django_filters.CharFilter(field_name='brand', lookup_expr='icontains', label="Brand (contains)")
    manufacturer = django_filters.CharFilter(field_name='manufacturer', lookup_expr='iexact', label="Manufacturer")
    
    # Price range filters
    price_min = django_filters.NumberFilter(field_name='base_price', lookup_expr='gte', label="Min price")
    price_max = django_filters.NumberFilter(field_name='base_price', lookup_expr='lte', label="Max price")
    
    # Stock filters
    in_stock = django_filters.BooleanFilter(method='filter_in_stock', label="In stock only")
    stock_min = django_filters.NumberFilter(field_name='stock_quantity', lookup_expr='gte', label="Min stock")
    stock_max = django_filters.NumberFilter(field_name='stock_quantity', lookup_expr='lte', label="Max stock")

    class Meta:
        model = Nomenklatura
        fields = [
            'code_1c', 'name', 'category', 'is_active',
            'description_status', 'image_status', 
            'project', 'project_id',
            'created_from', 'created_to', 
            'updated_from', 'updated_to'
        ]
    
    def filter_search(self, queryset, name, value):
        """Global search across multiple fields"""
        if value:
            return queryset.filter(
                Q(name__icontains=value) |
                Q(code_1c__icontains=value) |
                Q(article_code__icontains=value) |
                Q(brand__icontains=value) |
                Q(category__icontains=value)
            )
        return queryset
    
    def filter_in_stock(self, queryset, name, value):
        """Filter items in stock"""
        if value:
            return queryset.filter(stock_quantity__gt=0)
        return queryset

    def filter_description(self, queryset, name, value):
        if value == "with":
            return queryset.exclude(Q(description__isnull=True) | Q(description__exact=""))
        if value == "without":
            return queryset.filter(Q(description__isnull=True) | Q(description__exact=""))
        return queryset

    def filter_image_status(self, queryset, name, value):
        if value == "with":
            # Rasmlari bor bo'lganlar
            return queryset.filter(images__is_deleted=False).distinct()
        if value == "without":
            # Rasmlari yo'q bo'lganlar
            return queryset.exclude(images__is_deleted=False).distinct()
        return queryset


class NomenklaturaImageFilterSet(django_filters.FilterSet):
    created_from = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    created_to = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')

    project = django_filters.CharFilter(field_name='nomenklatura__project__code_1c', label="Project code_1c")
    project_id = django_filters.NumberFilter(field_name='nomenklatura__project__id', label="Project ID")
    code_1c = django_filters.CharFilter(field_name='nomenklatura__code_1c', label="Nomenklatura code_1c")
    article_code = django_filters.CharFilter(field_name='nomenklatura__article_code', label="Artikul kodi")

    class Meta:
        model = NomenklaturaImage
        fields = ['nomenklatura', 'code_1c', 'article_code', 'is_main', 'category', 'project', 'project_id', 'created_from', 'created_to']


from utils.mixins import ProjectScopedMixin

@extend_schema_view(
    list=extend_schema(
        tags=['Nomenklatura'],
        summary="Barcha nomenklatura ro'yxatini olish (Global)",
        description="Loyiha cheklovlarisiz barcha mavjud nomenklatura ro'yxati.",
    ),
    retrieve=extend_schema(
        tags=['Nomenklatura'],
        summary="Bitta nomenklatura ma'lumotini olish",
        description="`code_1c` identifikatoriga ko'ra nomenklatura ma'lumotlarini qaytaradi.",
    ),
    create=extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura yaratish",
        description="Yangi mahsulot qo'shish.",
        examples=[
            OpenApiExample(
                "Nomenklatura Create Example",
                value={
                    "code_1c": "NOM-001",
                    "name": "Coca-Cola 1.5L",
                    "category": "Drinks",
                    "base_price": 12000.00,
                    "sale_price": 13500.00,
                    "stock_quantity": 100,
                    "unit_of_measure": "dona",
                    "barcode": "478000000001",
                    "is_active": True
                },
                request_only=True
            )
        ]
    ),
    update=extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura ma'lumotlarini to'liq yangilash",
        examples=[
            OpenApiExample(
                "Nomenklatura Update Example",
                value={
                    "code_1c": "NOM-001",
                    "name": "Coca-Cola 1.5L (New Pack)",
                    "base_price": 12500.00,
                    "stock_quantity": 150,
                    "is_active": True
                },
                request_only=True
            )
        ]
    ),
    destroy=extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura soft-delete qilish",
        description="Mahsulotni o'chirmasdan, `is_deleted=True` qilib belgilaydi.",
    ),
)
class NomenklaturaViewSet(viewsets.ModelViewSet):
    """
    OPTIMIZED Nomenklatura ViewSet with Global Visibility
    """
    from utils.pagination import OptionalLimitOffsetPagination
    
    queryset = Nomenklatura.objects.filter(is_deleted=False)
    serializer_class = NomenklaturaSerializer
    pagination_class = OptionalLimitOffsetPagination
    lookup_field = 'code_1c'
    lookup_value_regex = '.+'
    filterset_class = NomenklaturaFilterSet
    search_fields = ['code_1c', 'article_code', 'name']
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        """Optimized queryset with global visibility"""
        queryset = super().get_queryset()
        queryset = queryset.prefetch_related(
            'images',
            'images__status',
            'images__source'
        ).order_by('-created_at')
        return queryset

    def perform_create(self, serializer):
        """Assign project if user has one, otherwise standard save"""
        user = self.request.user
        if not user.is_anonymous and hasattr(user, 'profile') and user.profile.project:
            from api.models import Project
            auth_project = user.profile.project
            api_project = Project.objects.filter(code_1c__iexact=auth_project.project_code, is_deleted=False).first()
            if api_project:
                serializer.save(project=api_project)
            else:
                serializer.save()
        else:
            serializer.save()
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    @action(detail=True, methods=['post'], url_path='enrich')
    def enrich(self, request, code_1c=None):
        """AI orqali bitta mahsulot ma'lumotlarini boyitish"""
        from nomenklatura.services import NomenklaturaEnrichmentService
        instance = self.get_object()
        service = NomenklaturaEnrichmentService()
        
        try:
            success = service.enrich_instance(instance)
            if success:
                return Response({'status': 'success', 'message': 'Mahsulot muvaffaqiyatli boyitildi'})
            else:
                return Response({'status': 'error', 'message': 'Boyitishda xatolik yuz berdi'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='bulk-enrich')
    def bulk_enrich(self, request):
        """AI orqali mahsulotlarni guruhlab boyitish"""
        from nomenklatura.services import NomenklaturaEnrichmentService
        code_1cs = request.data.get('code_1cs', [])
        project_id = request.data.get('project_id')
        
        queryset = self.get_queryset()
        if code_1cs:
            queryset = queryset.filter(code_1c__in=code_1cs)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
            
        service = NomenklaturaEnrichmentService()
        count = 0
        for instance in queryset:
            try:
                if service.enrich_instance(instance):
                    count += 1
            except:
                continue
                
        return Response({'status': 'success', 'message': f'{count} ta mahsulot boyitildi'})

    @action(detail=True, methods=['post'], url_path='clear-enrichment')
    def clear_enrichment(self, request, code_1c=None):
        """AI orqali qo'shilgan ma'lumotlarni tozalash"""
        from nomenklatura.services import NomenklaturaEnrichmentService
        instance = self.get_object()
        service = NomenklaturaEnrichmentService()
        
        try:
            service.clear_enrichment(instance)
            return Response({'status': 'success', 'message': "AI ma'lumotlari tozalandi"})
        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='bulk-clear')
    def bulk_clear(self, request):
        """AI ma'lumotlarini guruhlab tozalash"""
        from nomenklatura.services import NomenklaturaEnrichmentService
        code_1cs = request.data.get('code_1cs', [])
        project_id = request.data.get('project_id')
        
        queryset = self.get_queryset()
        if code_1cs:
            queryset = queryset.filter(code_1c__in=code_1cs)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
            
        service = NomenklaturaEnrichmentService()
        count = 0
        for instance in queryset:
            try:
                service.clear_enrichment(instance)
                count += 1
            except:
                continue
                
        return Response({'status': 'success', 'message': f'{count} ta mahsulot tozalandi'})

    # Excel helpers ---------------------------------------------------------
    @action(detail=False, methods=['get'])
    def duplicates(self, request):
        """
        Turli project'larda lekin bir xil code_1c ga ega nomenklaturalarni topish.
        """
        from django.db.models import Count
        
        # 1. code_1c bo'yicha guruhlash va soni > 1 bo'lganlarni topish
        duplicate_codes = Nomenklatura.objects.filter(is_deleted=False).values('code_1c').annotate(
            project_count=Count('project', distinct=True)
        ).filter(project_count__gt=1).values_list('code_1c', flat=True)
        
        # 2. Ushbu codelarga ega bo'lgan barcha objectlarni qaytarish
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(code_1c__in=duplicate_codes)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @staticmethod
    def _nomenklatura_excel_headers():
        return [
            'code_1c', 'name', 'title', 'description', 'is_active',
            'sku', 'barcode', 'brand', 'manufacturer', 'model', 'series', 'vendor_code',
            'base_price', 'sale_price', 'cost_price', 'currency', 'discount_percent', 'tax_rate',
            'stock_quantity', 'min_stock', 'max_stock', 'unit_of_measure', 'weight', 'dimensions', 'volume',
            'category', 'subcategory',
            'color', 'size', 'material', 'warranty_period', 'expiry_date', 'production_date',
            'notes', 'rating', 'popularity_score', 'seo_keywords', 'source'
        ]

    def _validate_nomenklatura_headers(self, sheet):
        expected = [header.lower() for header in self._nomenklatura_excel_headers()]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(expected), values_only=True), None)
        if not header_row:
            return False, expected, []
        normalized = [clean_cell(value).lower() for value in header_row[:len(expected)]]
        return normalized == expected, expected, normalized

    @extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura ma'lumotlarini Excel formatda eksport qilish",
        responses={200: OpenApiResponse(description="XLSX fayl")},
    )
    @action(detail=False, methods=['get'], url_path='export-xlsx', permission_classes=[IsAuthenticated])
    def export_xlsx(self, request):
        # Export uchun alohida queryset - prefetch_related kerak emas va SQLite limit muammosini oldini oladi
        base_queryset = Nomenklatura.objects.filter(is_deleted=False)
        queryset = self.filter_queryset(base_queryset)
        
        workbook = build_template_workbook('Nomenklatura', self._nomenklatura_excel_headers())
        sheet = workbook.active
        
        # iterator() ishlatish - memory-efficient va SQLite limit muammosini hal qiladi
        # Chunking bilan ishlash - har safar 1000 ta yozuvni qayta ishlash
        chunk_size = 1000
        offset = 0
        
        while True:
            chunk = queryset[offset:offset + chunk_size]
            if not chunk.exists():
                break
            
            for item in chunk.iterator():
                sheet.append([
                    item.code_1c,
                    item.name,
                    item.title or '',
                    item.description or '',
                    item.is_active,
                    item.sku or '',
                    item.barcode or '',
                    item.brand or '',
                    item.manufacturer or '',
                    item.model or '',
                    item.series or '',
                    item.vendor_code or '',
                    item.base_price or '',
                    item.sale_price or '',
                    item.cost_price or '',
                    item.currency or '',
                    item.discount_percent or '',
                    item.tax_rate or '',
                    item.stock_quantity or '',
                    item.min_stock or '',
                    item.max_stock or '',
                    item.unit_of_measure or '',
                    item.weight or '',
                    item.dimensions or '',
                    item.volume or '',
                    item.category or '',
                    item.subcategory or '',
                    item.color or '',
                    item.size or '',
                    item.material or '',
                    item.warranty_period or '',
                    item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else '',
                    item.production_date.strftime('%Y-%m-%d') if item.production_date else '',
                    item.notes or '',
                    item.rating or '',
                    item.popularity_score or '',
                    item.seo_keywords or '',
                    item.source or '',
                ])
            
            offset += chunk_size
        
        return workbook_to_response(workbook, 'nomenklatura.xlsx')

    @extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura Excel shablonini yuklab olish",
        responses={200: OpenApiResponse(description="XLSX shablon fayl")},
    )
    @action(detail=False, methods=['get'], url_path='template-xlsx', permission_classes=[IsAuthenticated])
    def template_xlsx(self, request):
        workbook = build_template_workbook(
            'Nomenklatura template',
            self._nomenklatura_excel_headers(),
            ['NOM-001', 'Namuna mahsulot', 'Sarlavha', '<p>Izoh</p>', True],
        )
        return workbook_to_response(workbook, 'nomenklatura_template.xlsx')

    @extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura ma'lumotlarini Excel fayldan import qilish",
        request={
            'multipart/form-data': inline_serializer(
                name='NomenklaturaImportPayload',
                fields={
                    'file': serializers.FileField(help_text='XLSX fayl'),
                    'project_id': serializers.IntegerField(required=False, help_text='Loyiha ID'),
                },
            )
        },
        responses={
            200: OpenApiResponse(description="Import natijalari (created/updated/errors)"),
            400: OpenApiResponse(description="Yaroqsiz fayl yoki header mos emas"),
        },
    )
    @action(
        detail=False,
        methods=['post'],
        url_path='import-xlsx',
        parser_classes=[MultiPartParser],
        permission_classes=[IsAuthenticated],
    )
    def import_xlsx(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'file field talab qilinadi'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(uploaded, data_only=True)
        except Exception:
            return Response({'error': 'XLSX faylni o\'qib bo\'lmadi'}, status=status.HTTP_400_BAD_REQUEST)

        sheet = workbook.active
        is_valid, expected, received = self._validate_nomenklatura_headers(sheet)
        if not is_valid:
            return Response(
                {'error': 'Excel headerlari mos emas', 'expected': expected, 'received': received},
                status=status.HTTP_400_BAD_REQUEST,
            )

        project_id = request.data.get('project_id')
        if project_id:
            try:
                project_id = int(str(project_id).strip())
            except (ValueError, TypeError):
                project_id = None
        
        # Create persistent log
        import_log = ImportLog.objects.create(
            entity_type='nomenklatura',
            filename=uploaded.name,
            status='processing',
            performed_by=request.user if request.user.is_authenticated else None
        )

        stats = {'created': 0, 'updated': 0, 'errors': [], 'errors_json': []}
        rows = list(sheet.iter_rows(min_row=2, max_col=len(expected), values_only=True))
        import_log.total_rows = len(rows)
        import_log.save()

        for idx, row in enumerate(rows, start=2):
            if not row or all(value in (None, '') for value in row):
                continue
            
            try:
                # Basic fields
                code = clean_cell(row[0])
                if not code:
                    error_msg = f"Row {idx}: code_1c bo'sh bo'lishi mumkin emas"
                    stats['errors'].append(error_msg)
                    continue
                
                name = clean_cell(row[1]) or code
                title = clean_cell(row[2])
                description = row[3] if row[3] is not None else ''
                is_active = parse_bool_cell(row[4], default=True)

                defaults = {
                    'name': name,
                    'title': title,
                    'description': description,
                    'is_active': is_active,
                    'is_deleted': False,
                    
                    # Extended fields based on _nomenklatura_excel_headers
                    'sku': clean_cell(row[5]),
                    'barcode': clean_cell(row[6]),
                    'brand': clean_cell(row[7]),
                    'manufacturer': clean_cell(row[8]),
                    'model': clean_cell(row[9]),
                    'series': clean_cell(row[10]),
                    'vendor_code': clean_cell(row[11]),
                    'base_price': row[12] if isinstance(row[12], (int, float)) else None,
                    'sale_price': row[13] if isinstance(row[13], (int, float)) else None,
                    'cost_price': row[14] if isinstance(row[14], (int, float)) else None,
                    'currency': clean_cell(row[15]) or 'UZS',
                    'discount_percent': row[16] if isinstance(row[16], (int, float)) else None,
                    'tax_rate': row[17] if isinstance(row[17], (int, float)) else None,
                    'stock_quantity': row[18] if isinstance(row[18], (int, float)) else None,
                    'min_stock': row[19] if isinstance(row[19], (int, float)) else None,
                    'max_stock': row[20] if isinstance(row[20], (int, float)) else None,
                    'unit_of_measure': clean_cell(row[21]),
                    'weight': row[22] if isinstance(row[22], (int, float)) else None,
                    'dimensions': clean_cell(row[23]),
                    'volume': row[24] if isinstance(row[24], (int, float)) else None,
                    'category': clean_cell(row[25]),
                    'subcategory': clean_cell(row[26]),
                    'color': clean_cell(row[27]),
                    'size': clean_cell(row[28]),
                    'material': clean_cell(row[29]),
                    'warranty_period': row[30] if isinstance(row[30], int) else None,
                    'notes': clean_cell(row[33]),
                    'rating': row[34] if isinstance(row[34], (int, float)) else None,
                    'popularity_score': row[35] if isinstance(row[35], int) else 0,
                    'seo_keywords': clean_cell(row[36]),
                    'source': clean_cell(row[37]),
                }
                
                # Date fields
                if row[31]: # expiry_date
                    defaults['expiry_date'] = row[31]
                if row[32]: # production_date
                    defaults['production_date'] = row[32]

                lookup = {'code_1c': code, 'is_deleted': False}
                if project_id:
                    lookup['project_id'] = project_id
                    defaults['project_id'] = project_id
                
                # Use filter().first() instead of update_or_create to avoid MultipleObjectsReturned
                # if there are duplicate codes across different projects or same project.
                obj = Nomenklatura.objects.filter(**lookup).first()
                if obj:
                    for key, value in defaults.items():
                        setattr(obj, key, value)
                    obj.save()
                    created_flag = False
                else:
                    obj = Nomenklatura.objects.create(**lookup, **defaults)
                    created_flag = True

                if created_flag:
                    stats['created'] += 1
                else:
                    stats['updated'] += 1
                    
            except Exception as exc:
                error_item = f"Row {idx} ({code if 'code' in locals() else 'Unknown'}): {str(exc)}"
                stats['errors'].append(error_item)
                stats['errors_json'].append({'row': idx, 'code': code if 'code' in locals() else '', 'error': str(exc)})

        # Finalize log
        import_log.created_count = stats['created']
        import_log.updated_count = stats['updated']
        import_log.error_count = len(stats['errors'])
        import_log.errors_json = stats.get('errors_json', [])
        import_log.status = 'completed' if not stats['errors'] else 'error'
        import_log.summary = f"Imported: {stats['created']}, Updated: {stats['updated']}, Errors: {len(stats['errors'])}"
        import_log.save()
        
        # Clear cache
        try:
            cache.delete_pattern("nomenklatura_*")
        except AttributeError:
            cache.clear()

        return Response(stats, status=status.HTTP_200_OK)

@extend_schema_view(
    list=extend_schema(
        tags=['Nomenklatura Image'],
        summary="Nomenklatura rasm ro'yxatini olish",
        description="""
Nomenklatura rasmlarini `nomenklatura` (ID) va `category` bo'yicha filterlash mumkin.

**is_main xususiyati:**
- Har bir nomenklatura uchun faqat **bitta** rasm `is_main=True` bo'lishi mumkin.
- Agar yangi rasm `is_main=True` qilib belgilansa, ushbu nomenklatura'ning boshqa barcha rasmlari avtomatik ravishda `is_main=False` holatiga o'tadi.
""",
        parameters=[
            OpenApiParameter(name="nomenklatura", type=OpenApiTypes.INT, description="Nomenklatura ID"),
            OpenApiParameter(name="project", type=OpenApiTypes.STR, description="Loyiha code_1c bo'yicha filtrlash"),
            OpenApiParameter(name="project_id", type=OpenApiTypes.INT, description="Loyiha ID bo'yicha filtrlash"),
            OpenApiParameter(
                name='is_main',
                required=False,
                type=OpenApiTypes.BOOL,
                description="Asosiy rasm bo'yicha filter",
            ),
            OpenApiParameter(
                name='category',
                required=False,
                type=OpenApiTypes.STR,
                description="Rasm toifasi bo'yicha filter",
            ),
            OpenApiParameter(
                name='created_from',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sanadan boshlab (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name='created_to',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sana chegarasi (YYYY-MM-DD)",
            ),
        ],
    ),
    create=extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura uchun bitta rasm yuklash",
        description="Multipart form-data orqali bitta rasm faylini yuklaydi.",
    ),
    destroy=extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura rasmini o'chirish",
        description="Rasmni bazadan o'chiradi yoki soft-delete qiladi.",
    ),
)
class NomenklaturaImageViewSet(viewsets.ModelViewSet):
    queryset = NomenklaturaImage.objects.filter(is_deleted=False)
    serializer_class = NomenklaturaImageSerializer
    filterset_class = NomenklaturaImageFilterSet
    search_fields = ['nomenklatura__code_1c', 'nomenklatura__article_code', 'nomenklatura__name']
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        """Optimizatsiya: bog'langan model'larni yuklash va global ko'rinish"""
        return super().get_queryset().select_related(
            'nomenklatura', 'nomenklatura__project', 'status', 'source'
        ).order_by('-created_at')
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    @extend_schema(
        tags=['Nomenklatura'],
        summary="Nomenklatura uchun bir nechta rasm yuklash",
        description=(
            "Multipart form-data formatida bir nechta rasmni birdaniga yuklaydi. "
            "`nomenklatura` maydoniga `code_1c` qiymati yuboriladi. Ixtiyoriy ravishda `category` "
            "va `note` maydonlari orqali rasmlarga umumiy teg yoki izoh qo'shish mumkin."
        ),
        request=NomenklaturaImageBulkUploadSerializer,
        responses={
            201: OpenApiResponse(
                response=NomenklaturaImageSerializer(many=True),
                description="Yangi yuklangan rasmlar ro'yxati",
            ),
            400: OpenApiResponse(description="Kerakli maydonlar yetishmaydi"),
            404: OpenApiResponse(description="Nomenklatura topilmadi"),
        },
        examples=[
            OpenApiExample(
                name="Multipart sample",
                description="HTTPie yordamida bir nechta nomenklatura rasmini yuborish",
                value="http --form POST /api/v1/nomenklatura-images/bulk-upload/ nomenklatura=N-001 "
                "category=kir yuvish note='Telegram katalogi uchun' "
                "images@/path/img1.jpg images@/path/img2.jpg",
            )
        ],
    )
    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        """Bir vaqtda bir nechta rasm yuklash"""
        nomenklatura_code = request.data.get('nomenklatura')
        if not nomenklatura_code:
            return Response(
                {'error': 'nomenklatura code_1c talab qilinadi'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            project_id = request.data.get('project_id')
            query = Q(code_1c=nomenklatura_code, is_deleted=False)
            if project_id:
                try:
                    project_id = int(str(project_id).strip())
                    query &= Q(project_id=project_id)
                except (ValueError, TypeError):
                    pass
            
            # Agar project_id berilmagan bo'lsa va MultipleObjectsReturned bo'lsa, xatolik qaytaramiz
            nomenklatura = Nomenklatura.objects.get(query)
        except Nomenklatura.DoesNotExist:
            return Response(
                {'error': 'Nomenklatura topilmadi'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Nomenklatura.MultipleObjectsReturned:
            return Response(
                {'error': 'Ushbu kod bir nechta loyihada mavjud. Iltimos, project_id ni yuboring.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        
        images = request.FILES.getlist('images')
        if not images:
            return Response(
                {'error': 'Rasmlar talab qilinadi'},
                status=status.HTTP_400_BAD_REQUEST
            )

        category = request.data.get('category', '')
        note = request.data.get('note', '')
        
        created_images = []
        for image in images:
            image_obj = NomenklaturaImage.objects.create(
                nomenklatura=nomenklatura,
                image=image,
                is_main=False,
                is_active=True,
                is_deleted=False,
                category=category,
                note=note,
            )
            serializer = NomenklaturaImageSerializer(image_obj, context={'request': request})
            created_images.append(serializer.data)
        
        return Response({
            'message': f'{len(created_images)} ta rasm muvaffaqiyatli yuklandi',
            'images': created_images
        }, status=status.HTTP_201_CREATED)
