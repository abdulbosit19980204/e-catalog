import django_filters
from django.db.models import Q
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from drf_spectacular.types import OpenApiTypes
from django.core.cache import cache
from utils.cache import smart_cache_get, smart_cache_set, smart_cache_delete
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
    extend_schema_view,
    inline_serializer,
)
from openpyxl import load_workbook
from utils.excel import build_template_workbook, workbook_to_response, parse_bool_cell, clean_cell
from .models import Client, ClientImage
from .serializers import ClientImageBulkUploadSerializer, ClientImageSerializer, ClientSerializer


class ClientFilterSet(django_filters.FilterSet):
    """Enhanced filtering with backward compatibility"""
    # EXISTING FILTERS (preserved)
    description_status = django_filters.ChoiceFilter(
        label="Description status",
        method="filter_description",
        choices=(("with", "Description bor"), ("without", "Description yo'q")),
    )
    image_status = django_filters.ChoiceFilter(
        label="Image status",
        method="filter_image_status",
        choices=(("with", "Rasm bor"), ("without", "Rasm yo'q")),
    )
    project = django_filters.CharFilter(field_name='project__code_1c', label="Project code_1c")
    project_id = django_filters.NumberFilter(field_name='project__id', label="Project ID")
    created_from = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    created_to = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')
    updated_from = django_filters.DateFilter(field_name='updated_at', lookup_expr='date__gte')
    updated_to = django_filters.DateFilter(field_name='updated_at', lookup_expr='date__lte')
    
    # NEW FILTERS (additive)
    search = django_filters.CharFilter(method='filter_search', label="Global search")
    name_contains = django_filters.CharFilter(field_name='name', lookup_expr='icontains')
    email_contains = django_filters.CharFilter(field_name='email', lookup_expr='icontains')
    code_contains = django_filters.CharFilter(field_name='client_code_1c', lookup_expr='icontains')

    class Meta:
        model = Client
        fields = [
            'client_code_1c', 'name', 'email', 'project', 'project_id',
            'created_from', 'created_to', 'updated_from', 'updated_to', 'is_active'
        ]
    
    def filter_search(self, queryset, name, value):
        """Global search across multiple fields"""
        if value:
            return queryset.filter(
                Q(name__icontains=value) |
                Q(client_code_1c__icontains=value) |
                Q(email__icontains=value) |
                Q(phone__icontains=value) |
                Q(tax_id__icontains=value)
            )
        return queryset

    def filter_description(self, queryset, name, value):
        if value == "with":
            return queryset.exclude(Q(description__isnull=True) | Q(description__exact=""))
        if value == "without":
            return queryset.filter(Q(description__isnull=True) | Q(description__exact=""))
        return queryset

    def filter_image_status(self, queryset, name, value):
        if value == "with":
            # Rasmlari bor bo'lganlar
            return queryset.filter(images__is_deleted=False).distinct()
        if value == "without":
            # Rasmlari yo'q bo'lganlar
            return queryset.exclude(images__is_deleted=False).distinct()
        return queryset


class ClientImageFilterSet(django_filters.FilterSet):
    client_code_1c = django_filters.CharFilter(field_name='client__client_code_1c')
    created_from = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    created_to = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')

    class Meta:
        model = ClientImage
        fields = ['client', 'client_code_1c', 'status', 'is_main', 'category', 'created_from', 'created_to']


from utils.mixins import ProjectScopedMixin

@extend_schema_view(
    list=extend_schema(
        tags=['Clients'],
        summary="Client ro'yxatini olish",
        description=(
            "Authentication talab qilinadi. Aktiv va soft-delete qilinmagan (is_deleted=False) client'lar ro'yxatini pagination bilan qaytaradi. "
            " `search` parametri orqali `name`, `email` va `client_code_1c` bo'yicha qidirish mumkin. "
            "Ma'muotlar ProjectScopedMixin orqali foydalanuvchi proyektiga ko'ra filtrlanadi."
        ),
        parameters=[
            OpenApiParameter(
                name='search',
                required=False,
                type=OpenApiTypes.STR,
                description="Client nomi, email yoki code bo'yicha qidirish",
            ),
        ],
    ),
    retrieve=extend_schema(
        tags=['Clients'],
        summary="Bitta client ma'lumotini olish",
        description="`client_code_1c` identifikatoriga ko'ra client ma'lumotlarini qaytaradi.",
    ),
)
class ClientViewSet(ProjectScopedMixin, viewsets.ModelViewSet):
    """
    OPTIMIZED Client ViewSet with Project Isolation
    """
    from utils.pagination import OptionalLimitOffsetPagination
    
    queryset = Client.objects.filter(is_deleted=False)
    serializer_class = ClientSerializer
    pagination_class = OptionalLimitOffsetPagination
    lookup_field = "client_code_1c"
    lookup_value_regex = "[^/]+"
    filterset_class = ClientFilterSet
    search_fields = ["name", "email", "client_code_1c"]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Optimized queryset with multi-tenant isolation"""
        # Mixin filters by project
        queryset = super().get_queryset()
        queryset = queryset.prefetch_related(
            'images',
            'images__status',
            'images__source'
        ).order_by('-created_at')
        return queryset
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    # Excel helpers ---------------------------------------------------------
    @action(detail=False, methods=['get'])
    def duplicates(self, request):
        """
        Turli project'larda lekin bir xil code_1c ga ega clientlarni topish.
        """
        from django.db.models import Count
        
        # 1. code_1c bo'yicha guruhlash va soni > 1 bo'lganlarni topish
        duplicate_codes = Client.objects.filter(is_deleted=False).values('client_code_1c').annotate(
            project_count=Count('project', distinct=True)
        ).filter(project_count__gt=1).values_list('client_code_1c', flat=True)
        
        # 2. Ushbu codelarga ega bo'lgan barcha objectlarni qaytarish
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(client_code_1c__in=duplicate_codes)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @staticmethod
    def _client_excel_headers():
        return [
            'client_code_1c', 'name', 'email', 'phone', 'description', 'is_active',
            'company_name', 'tax_id', 'registration_number', 'legal_address', 'actual_address',
            'fax', 'website', 'industry', 'business_type', 'employee_count', 'annual_revenue',
            'established_date', 'payment_terms', 'credit_limit', 'currency',
            'city', 'region', 'country', 'postal_code',
            'contact_person', 'contact_position', 'contact_email', 'contact_phone',
            'notes', 'rating', 'priority', 'source'
        ]

    def _validate_client_headers(self, sheet):
        expected = [header.lower() for header in self._client_excel_headers()]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(expected), values_only=True), None)
        if not header_row:
            return False, expected, []
        normalized = [clean_cell(value).lower() for value in header_row[:len(expected)]]
        return normalized == expected, expected, normalized

    @extend_schema(
        tags=['Clients'],
        summary="Client ma'lumotlarini Excel formatda eksport qilish",
        responses={200: OpenApiResponse(description="XLSX fayl")},
    )
    @action(detail=False, methods=['get'], url_path='export-xlsx', permission_classes=[IsAuthenticated])
    def export_xlsx(self, request):
        # Export uchun alohida queryset - prefetch_related kerak emas va SQLite limit muammosini oldini oladi
        base_queryset = Client.objects.filter(is_deleted=False)
        queryset = self.filter_queryset(base_queryset)
        
        workbook = build_template_workbook('Clients', self._client_excel_headers())
        sheet = workbook.active
        
        # iterator() ishlatish - memory-efficient va SQLite limit muammosini hal qiladi
        # Chunking bilan ishlash - har safar 1000 ta yozuvni qayta ishlash
        chunk_size = 1000
        offset = 0
        
        while True:
            chunk = queryset[offset:offset + chunk_size]
            if not chunk.exists():
                break
            
            for client in chunk.iterator():
                sheet.append([
                    client.client_code_1c,
                    client.name,
                    client.email or '',
                    client.phone or '',
                    client.description or '',
                    client.is_active,
                    client.company_name or '',
                    client.tax_id or '',
                    client.registration_number or '',
                    client.legal_address or '',
                    client.actual_address or '',
                    client.fax or '',
                    client.website or '',
                    client.industry or '',
                    client.business_type or '',
                    client.employee_count or '',
                    client.annual_revenue or '',
                    client.established_date.strftime('%Y-%m-%d') if client.established_date else '',
                    client.payment_terms or '',
                    client.credit_limit or '',
                    client.currency or '',
                    client.city or '',
                    client.region or '',
                    client.country or '',
                    client.postal_code or '',
                    client.contact_person or '',
                    client.contact_position or '',
                    client.contact_email or '',
                    client.contact_phone or '',
                    client.notes or '',
                    client.rating or '',
                    client.priority or '',
                    client.source or '',
                ])
            
            offset += chunk_size
        
        return workbook_to_response(workbook, 'clients.xlsx')

    @extend_schema(
        tags=['Clients'],
        summary="Client Excel shablonini yuklab olish",
        responses={200: OpenApiResponse(description="XLSX shablon fayl")},
    )
    @action(detail=False, methods=['get'], url_path='template-xlsx', permission_classes=[IsAuthenticated])
    def template_xlsx(self, request):
        workbook = build_template_workbook(
            'Clients template',
            self._client_excel_headers(),
            ['CLI-001', 'Namuna client', 'client@example.com', '+998 90 000 00 00', 'Izoh', True],
        )
        return workbook_to_response(workbook, 'clients_template.xlsx')

    @extend_schema(
        tags=['Clients'],
        summary="Client ma'lumotlarini Excel fayldan import qilish",
        request={
            'multipart/form-data': inline_serializer(
                name='ClientImportPayload',
                fields={
                    'file': serializers.FileField(help_text='XLSX fayl'),
                },
            )
        },
        responses={
            200: OpenApiResponse(description="Import natijalari (created/updated/errors)"),
            400: OpenApiResponse(description="Yaroqsiz fayl yoki header mos emas"),
        },
    )
    @action(
        detail=False,
        methods=['post'],
        url_path='import-xlsx',
        parser_classes=[MultiPartParser],
        permission_classes=[IsAuthenticated],
    )
    def import_xlsx(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'file field talab qilinadi'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(uploaded, data_only=True)
        except Exception:
            return Response({'error': 'XLSX faylni o\'qib bo\'lmadi'}, status=status.HTTP_400_BAD_REQUEST)

        sheet = workbook.active
        is_valid, expected, received = self._validate_client_headers(sheet)
        if not is_valid:
            return Response(
                {'error': 'Excel headerlari mos emas', 'expected': expected, 'received': received},
                status=status.HTTP_400_BAD_REQUEST,
            )

        project_id = request.data.get('project_id')
        if project_id:
            try:
                project_id = int(str(project_id).strip())
            except (ValueError, TypeError):
                project_id = None
        
        stats = {'created': 0, 'updated': 0, 'errors': []}
        for idx, row in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(expected), values_only=True),
            start=2,
        ):
            if not row or all(value in (None, '') for value in row):
                continue
            code = clean_cell(row[0])
            if not code:
                stats['errors'].append(f"Row {idx}: client_code_1c bo'sh bo'lishi mumkin emas")
                continue
            name = clean_cell(row[1]) or code
            email = clean_cell(row[2]) or None
            phone = clean_cell(row[3]) or None
            description = row[4] if row[4] is not None else ''
            is_active = parse_bool_cell(row[5], default=True)

            defaults = {
                'name': name,
                'email': email or None,
                'phone': phone or None,
                'description': description,
                'is_active': is_active,
                'is_deleted': False,
            }
            try:
                lookup = {'client_code_1c': code, 'is_deleted': False}
                if project_id:
                    lookup['project_id'] = project_id
                    defaults['project_id'] = project_id
                
                obj, created_flag = Client.objects.update_or_create(
                    **lookup,
                    defaults=defaults,
                )

                if created_flag:
                    stats['created'] += 1
                else:
                    stats['updated'] += 1
            except Exception as exc:  # noqa: BLE001
                stats['errors'].append(f"Row {idx}: {exc}")
        return Response(stats, status=status.HTTP_200_OK)

@extend_schema_view(
    list=extend_schema(
        tags=['Clients'],
        summary="Client rasmlari ro'yxatini olish",
        description="""
Client rasmlarini `client` (ID) va `client_code_1c` bo'yicha filterlash mumkin.

**is_main xususiyati:**
- Har bir client uchun faqat **bitta** rasm `is_main=True` bo'lishi mumkin.
- Agar yangi rasm `is_main=True` qilib belgilansa, ushbu client'ning boshqa barcha rasmlari avtomatik ravishda `is_main=False` holatiga o'tadi.
""",
        parameters=[
            OpenApiParameter(
                name='client',
                required=False,
                type=OpenApiTypes.INT,
                description="Client ID bo'yicha filter",
            ),
            OpenApiParameter(
                name='client_code_1c',
                required=False,
                type=OpenApiTypes.STR,
                description="Client client_code_1c bo'yicha filter",
            ),
            OpenApiParameter(
                name='is_main',
                required=False,
                type=OpenApiTypes.BOOL,
                description="Asosiy rasm bo'yicha filter",
            ),
            OpenApiParameter(
                name='category',
                required=False,
                type=OpenApiTypes.STR,
                description="Rasm toifasi bo'yicha filter",
            ),
            OpenApiParameter(
                name='created_from',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sanadan boshlab (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name='created_to',
                required=False,
                type=OpenApiTypes.DATE,
                description="Yaratilgan sana chegarasi (YYYY-MM-DD)",
            ),
        ],
    ),
    create=extend_schema(
        tags=['Clients'],
        summary="Client uchun bitta rasm yuklash",
        description="Multipart form-data orqali bitta rasm faylini yuklaydi.",
    ),
    destroy=extend_schema(
        tags=['Clients'],
        summary="Client rasmini o'chirish",
        description="Rasmni bazadan o'chiradi yoki soft-delete qiladi.",
    ),
)
class ClientImageViewSet(ProjectScopedMixin, viewsets.ModelViewSet):
    queryset = ClientImage.objects.filter(is_deleted=False)
    serializer_class = ClientImageSerializer
    filterset_class = ClientImageFilterSet
    search_fields = ['client__client_code_1c', 'client__name']
    permission_classes = [IsAuthenticated]  # Faqat authenticated user'lar uchun
    
    def perform_create(self, serializer):
        super().perform_create(serializer)
        cache.clear()

    def perform_update(self, serializer):
        super().perform_update(serializer)
        cache.clear()

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.save(update_fields=['is_deleted', 'updated_at'])
        cache.clear()
    
    @extend_schema(
        tags=['Clients'],
        summary="Client uchun ko'plab rasm yuklash",
        request=ClientImageBulkUploadSerializer,
        responses={201: ClientImageSerializer(many=True)},
    )
    @action(detail=False, methods=['post'], url_path='bulk-upload', parser_classes=[MultiPartParser])
    def bulk_upload(self, request):
        serializer = ClientImageBulkUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        client_obj = serializer.validated_data['client']
        images = serializer.validated_data['images']
        category = serializer.validated_data.get('category', 'other')
        note = serializer.validated_data.get('note', '')
        
        created_images = []
        for img_file in images:
            image_obj = ClientImage.objects.create(
                client=client_obj,
                image=img_file,
                category=category,
                note=note
            )
            created_images.append(image_obj)
            
        cache.clear()
        return Response(
            ClientImageSerializer(created_images, many=True, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )
    
    def get_queryset(self):
        """Optimizatsiya: bog'langan model'larni yuklash va multi-tenant isolation"""
        # Mixin filters by project
        queryset = super().get_queryset()
        queryset = queryset.select_related('client', 'client__project', 'status', 'source').order_by('-created_at')
        return queryset

@extend_schema_view(
    list=extend_schema(
        tags=['Agent Visits (Legacy)'],
        summary="Agent tashrifi rasmlari ro'yxatini olish",
        description=(
            "Agentlar tomonidan magazinlarga tashrif buyurilgan (visit) vaqtida olingan barcha rasmlarni qaytaradi. "
            "Ma'muotlar ProjectScopedMixin orqali foydalanuvchi proyektiga ko'ra filtrlanadi."
        ),
    ),
    retrieve=extend_schema(tags=['Agent Visits (Legacy)'], summary="Bitta tashrif rasmi tafsiloti"),
)
class VisitImageViewSet(ProjectScopedMixin, viewsets.ModelViewSet):
    """
    Legacy Agent visits viewset with project isolation
    """
    queryset = ClientImage.objects.filter(is_deleted=False)
    serializer_class = ClientImageSerializer
    permission_classes = [IsAuthenticated]
    filterset_class = ClientImageFilterSet
    search_fields = ['client__name', 'client__client_code_1c', 'note']

    def get_queryset(self):
        # Mixin filters by project
        return super().get_queryset().select_related(
            'client', 'client__project', 'status', 'source'
        ).only(
            'id', 'client__id', 'client__name', 'client__client_code_1c', 'client__project',
            'image', 'is_main', 'category', 'note', 'status', 'source',
            'created_at'
        ).order_by('-created_at')

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
