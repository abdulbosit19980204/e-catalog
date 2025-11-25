from io import BytesIO
from typing import Iterable, List, Sequence

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def workbook_to_response(workbook: Workbook, filename: str) -> HttpResponse:
    """Serialize a workbook to an HttpResponse attachment."""
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type=EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_template_workbook(
    title: str,
    headers: Sequence[str],
    sample_row: Sequence[str] | None = None,
) -> Workbook:
    """Create a simple workbook containing only headers (and optional sample row)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(list(headers))
    if sample_row:
        sheet.append(list(sample_row))

    for idx, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 5, 15)
        sheet.column_dimensions[get_column_letter(idx)].width = width
    return workbook


def parse_bool_cell(value, default: bool = True) -> bool:
    """Convert Excel cell values to boolean."""
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    normalized = str(value).strip().lower()
    if not normalized:
        return default
    return normalized in {"1", "true", "yes", "ha", "active", "on", "y"}


def clean_cell(value) -> str:
    """Normalize cell value to trimmed string."""
    if value is None:
        return ""
    return str(value).strip()

